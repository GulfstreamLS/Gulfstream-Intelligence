import openpyxl
import os

file_path = "fda_metadata.xlsx"
if not os.path.exists(file_path):
    print(f"File {file_path} not found")
    exit(1)

wb = openpyxl.load_workbook(file_path)
sheet = wb.active

print(f"Sheet title: {sheet.title}")
print(f"Max rows: {sheet.max_row}")

# Check first 20 rows
for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20)):
    row_values = []
    for cell in row:
        val = cell.value
        link = cell.hyperlink.target if cell.hyperlink else None
        if link:
            row_values.append(f"LINK:{link}")
        else:
            row_values.append(val)
    print(f"Row {i}: {row_values}")
